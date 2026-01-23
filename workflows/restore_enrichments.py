#!/usr/bin/env python3
"""
Restore enriched hotel data from S3 export.

Reads FL_dbpr.xlsx from S3 and restores:
- Website
- Email
- Booking Engine
- Room Count

Matches hotels by name + city using batch operations.

Usage:
    uv run python -m workflows.restore_enrichments
    uv run python -m workflows.restore_enrichments --dry-run
"""

import argparse
import asyncio
import sys
import os

from loguru import logger
from openpyxl import load_workbook

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from db.client import init_db, get_conn

S3_PATH = "s3://sadie-gtm/HotelLeadGen/USA/FL/FL_dbpr.xlsx"
LOCAL_PATH = "/tmp/FL_dbpr.xlsx"


async def download_from_s3():
    """Download the xlsx file from S3."""
    import subprocess
    logger.info(f"Downloading {S3_PATH}...")
    result = subprocess.run(["aws", "s3", "cp", S3_PATH, LOCAL_PATH], capture_output=True)
    if result.returncode != 0:
        raise Exception(f"Failed to download: {result.stderr.decode()}")
    logger.info(f"Downloaded to {LOCAL_PATH}")


def load_xlsx():
    """Load and parse the xlsx file."""
    wb = load_workbook(LOCAL_PATH, read_only=True)
    ws = wb.active

    headers = [cell.value for cell in ws[1]]
    logger.info(f"Headers: {headers}")

    records = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        record = dict(zip(headers, row))
        records.append(record)

    logger.info(f"Loaded {len(records)} records from xlsx")
    return records


async def restore_enrichments(records, dry_run=False):
    """Restore enrichments to database using batch operations."""
    await init_db()

    async with get_conn() as conn:
        # Load all FL hotels into memory for matching
        logger.info("Loading FL hotels from database...")
        hotels = await conn.fetch("""
            SELECT id, UPPER(name) as name, UPPER(city) as city, website, email
            FROM sadie_gtm.hotels WHERE state = 'FL'
        """)
        hotel_map = {(r['name'], r['city']): r for r in hotels}
        logger.info(f"Loaded {len(hotel_map)} FL hotels")

        # Load booking engine map
        be_rows = await conn.fetch("SELECT id, LOWER(name) as name FROM sadie_gtm.booking_engines")
        be_map = {r['name']: r['id'] for r in be_rows}
        logger.info(f"Loaded {len(be_map)} booking engines")

        # Match records
        matched = []
        not_found = 0
        for record in records:
            name = record.get('Hotel')
            city = record.get('City')
            if not name or not city:
                continue

            key = (name.upper(), city.upper())
            hotel = hotel_map.get(key)
            if hotel:
                matched.append((hotel, record))
            else:
                not_found += 1

        logger.info(f"Matched {len(matched)} hotels, {not_found} not found")

        if dry_run:
            return {'matched': len(matched), 'not_found': not_found}

        # Batch update websites
        website_updates = [
            (r['Website'], h['id'])
            for h, r in matched
            if r.get('Website') and not h['website']
        ]
        if website_updates:
            await conn.executemany(
                "UPDATE sadie_gtm.hotels SET website = $1, updated_at = NOW() WHERE id = $2",
                website_updates
            )
            logger.info(f"Updated {len(website_updates)} websites")

        # Batch update emails
        email_updates = [
            (r['Email'], h['id'])
            for h, r in matched
            if r.get('Email') and not h['email']
        ]
        if email_updates:
            await conn.executemany(
                "UPDATE sadie_gtm.hotels SET email = $1, updated_at = NOW() WHERE id = $2",
                email_updates
            )
            logger.info(f"Updated {len(email_updates)} emails")

        # Batch insert booking engines
        be_inserts = []
        for h, r in matched:
            be_name = r.get('Booking Engine')
            if be_name and be_name.lower() in be_map:
                be_inserts.append((h['id'], be_map[be_name.lower()]))

        if be_inserts:
            await conn.executemany("""
                INSERT INTO sadie_gtm.hotel_booking_engines (hotel_id, booking_engine_id, detection_method, status)
                VALUES ($1, $2, 'restored_from_export', 1)
                ON CONFLICT (hotel_id) DO NOTHING
            """, be_inserts)
            logger.info(f"Inserted {len(be_inserts)} booking engine records")

        # Batch insert room counts
        rc_inserts = []
        for h, r in matched:
            rc = r.get('Room Count')
            if rc:
                try:
                    rc_inserts.append((h['id'], int(rc)))
                except (ValueError, TypeError):
                    pass

        if rc_inserts:
            await conn.executemany("""
                INSERT INTO sadie_gtm.hotel_room_count (hotel_id, room_count, source, status)
                VALUES ($1, $2, 'restored_from_export', 1)
                ON CONFLICT (hotel_id) DO UPDATE SET room_count = EXCLUDED.room_count
            """, rc_inserts)
            logger.info(f"Inserted {len(rc_inserts)} room count records")

    return {
        'matched': len(matched),
        'not_found': not_found,
        'website_updated': len(website_updates),
        'email_updated': len(email_updates),
        'booking_engine_updated': len(be_inserts),
        'room_count_updated': len(rc_inserts),
    }


async def main():
    parser = argparse.ArgumentParser(description="Restore enriched data from S3 export")
    parser.add_argument("--dry-run", action="store_true", help="Match only, don't update")
    parser.add_argument("--skip-download", action="store_true", help="Skip S3 download, use local file")
    args = parser.parse_args()

    logger.remove()
    logger.add(sys.stderr, level="INFO", format="<level>{level: <8}</level> | {message}")

    if not args.skip_download:
        await download_from_s3()

    records = load_xlsx()

    logger.info(f"{'DRY RUN - ' if args.dry_run else ''}Restoring enrichments...")
    stats = await restore_enrichments(records, dry_run=args.dry_run)

    logger.info("")
    logger.info("=" * 60)
    logger.info("Restore Complete" + (" (DRY RUN)" if args.dry_run else ""))
    logger.info("=" * 60)
    logger.info(f"Hotels matched: {stats['matched']}")
    logger.info(f"Hotels not found: {stats['not_found']}")
    if not args.dry_run:
        logger.info(f"Websites updated: {stats.get('website_updated', 0)}")
        logger.info(f"Emails updated: {stats.get('email_updated', 0)}")
        logger.info(f"Booking engines updated: {stats.get('booking_engine_updated', 0)}")
        logger.info(f"Room counts updated: {stats.get('room_count_updated', 0)}")


if __name__ == "__main__":
    asyncio.run(main())
