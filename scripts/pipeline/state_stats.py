#!/usr/bin/env python3
"""
Sadie State Stats - Generate Aggregate Stats per State
=======================================================
Scans OneDrive state folders and creates state-level aggregate stats.

Usage:
    python3 scripts/pipeline/state_stats.py --country usa
    python3 scripts/pipeline/state_stats.py --country usa --state florida
"""

import os
import sys
import argparse
import re
from pathlib import Path
from datetime import datetime
from collections import defaultdict

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    print("Error: openpyxl not installed. Install with: pip3 install openpyxl")
    sys.exit(1)

# OneDrive paths
ONEDRIVE_BASE = os.path.expanduser("~/Library/CloudStorage/OneDrive-ValsoftCorporation")
SADIE_FOLDER = "Sadie Lead Gen"

# Top 25 priority cities for Florida
TOP_25_FLORIDA_CITIES = {
    "miami beach", "kissimmee", "miami", "pensacola", "fort lauderdale",
    "tampa", "saint augustine", "st augustine", "key west", "windermere",
    "panama city beach", "bay pines", "orlando", "daytona beach",
    "north miami beach", "pompano beach", "homestead", "fort myers beach",
    "hialeah", "saint petersburg", "st petersburg", "clearwater beach",
    "jacksonville", "sarasota", "pembroke pines", "fort myers", "high springs"
}


def get_sadie_path() -> Path:
    return Path(ONEDRIVE_BASE) / SADIE_FOLDER


def find_city_files(state_path: Path) -> list[Path]:
    """Find all Excel files in a state folder (not recursive, direct children only)."""
    files = []
    for xlsx in state_path.glob("*.xlsx"):
        # Skip stats files
        if "stats" in xlsx.name.lower():
            continue
        # Skip temp files
        if xlsx.name.startswith("~"):
            continue
        files.append(xlsx)
    return files


def extract_stats_from_excel(xlsx_path: Path) -> dict:
    """Extract stats from a city Excel file's Stats sheet."""
    stats = {
        "city": xlsx_path.stem,  # filename without extension
        "file": xlsx_path.name,
        "leads_count": 0,
        "hotels_scraped": 0,
        "with_website": 0,
        "booking_found": 0,
        "tier1": 0,
        "tier2": 0,
        "with_phone": 0,
        "with_email": 0,
        "booking_engines": defaultdict(int),
    }

    try:
        wb = load_workbook(xlsx_path, read_only=True, data_only=True)

        # Get leads count from Leads sheet
        if "Leads" in wb.sheetnames:
            ws_leads = wb["Leads"]
            stats["leads_count"] = ws_leads.max_row - 1  # Subtract header
            stats["booking_found"] = stats["leads_count"]  # Default if no stats sheet

            # Find booking_engine column
            header_row = list(ws_leads.iter_rows(min_row=1, max_row=1, values_only=True))[0]
            be_col = None
            for i, h in enumerate(header_row):
                if h and str(h).lower() == "booking_engine":
                    be_col = i
                    break

            # Count booking engines
            if be_col is not None:
                for row in ws_leads.iter_rows(min_row=2, values_only=True):
                    if row and len(row) > be_col and row[be_col]:
                        engine = str(row[be_col]).strip().lower()
                        if engine and engine != "none":
                            stats["booking_engines"][engine] += 1

        # Try to parse Stats sheet for metrics
        if "Stats" in wb.sheetnames:
            ws_stats = wb["Stats"]

            # Read all cells into a dict for easier parsing
            cells = {}
            for row in ws_stats.iter_rows(max_row=30, max_col=6):
                for cell in row:
                    if cell.value:
                        cells[(cell.row, cell.column)] = str(cell.value)

            def extract_num(s):
                if not s:
                    return 0
                match = re.search(r'[\d,]+', s.replace(',', ''))
                if match:
                    try:
                        return int(match.group().replace(',', ''))
                    except:
                        pass
                return 0

            for (row, col), value in cells.items():
                val_lower = value.lower()
                next_val = cells.get((row, col + 1), "")

                if "hotels scraped" in val_lower:
                    stats["hotels_scraped"] = extract_num(next_val)
                elif "with website" in val_lower and "rate" not in val_lower:
                    stats["with_website"] = extract_num(next_val)
                elif "booking found" in val_lower:
                    stats["booking_found"] = extract_num(next_val)
                elif "tier 1" in val_lower:
                    stats["tier1"] = extract_num(next_val)
                elif "tier 2" in val_lower:
                    stats["tier2"] = extract_num(next_val)
                elif "with phone" in val_lower:
                    stats["with_phone"] = extract_num(next_val)
                elif "with email" in val_lower:
                    stats["with_email"] = extract_num(next_val)

        wb.close()

    except Exception as e:
        print(f"    Warning: Could not read {xlsx_path.name}: {e}")

    return stats


def generate_state_stats(state_path: Path, city_stats: list[dict]) -> Path:
    """Generate state-level aggregate stats Excel file."""

    state_name = state_path.name

    # Aggregate stats
    totals = {
        "hotels_scraped": sum(c["hotels_scraped"] for c in city_stats),
        "with_website": sum(c["with_website"] for c in city_stats),
        "booking_found": sum(c["booking_found"] for c in city_stats),
        "leads_count": sum(c["leads_count"] for c in city_stats),
        "tier1": sum(c["tier1"] for c in city_stats),
        "tier2": sum(c["tier2"] for c in city_stats),
        "with_phone": sum(c["with_phone"] for c in city_stats),
        "with_email": sum(c["with_email"] for c in city_stats),
    }

    # Aggregate booking engines
    all_engines = defaultdict(int)
    for c in city_stats:
        for engine, count in c["booking_engines"].items():
            all_engines[engine] += count

    # Calculate rates
    website_rate = round(totals["with_website"] / totals["hotels_scraped"] * 100, 1) if totals["hotels_scraped"] else 0
    detection_rate = round(totals["booking_found"] / totals["with_website"] * 100, 1) if totals["with_website"] else 0
    tier1_rate = round(totals["tier1"] / totals["booking_found"] * 100, 1) if totals["booking_found"] else 0

    # Create workbook
    wb = Workbook()

    # Styles
    title_font = Font(bold=True, size=16, color="FFFFFF")
    title_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
    section_font = Font(bold=True, size=11, color="FFFFFF")
    section_fill = PatternFill(start_color="455A64", end_color="455A64", fill_type="solid")
    label_font = Font(size=10)
    value_font = Font(bold=True, size=12)
    green_fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
    center = Alignment(horizontal="center", vertical="center")
    right = Alignment(horizontal="right", vertical="center")
    thin_border = Border(
        left=Side(style='thin', color='BDBDBD'),
        right=Side(style='thin', color='BDBDBD'),
        top=Side(style='thin', color='BDBDBD'),
        bottom=Side(style='thin', color='BDBDBD')
    )

    ws = wb.active
    ws.title = "State Stats"

    # Column widths
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 3
    ws.column_dimensions["D"].width = 25
    ws.column_dimensions["E"].width = 18

    # Title
    ws.merge_cells('A1:E1')
    c = ws.cell(row=1, column=1, value=f"{state_name.upper()} — STATE STATS")
    c.font = title_font
    c.fill = title_fill
    c.alignment = center
    ws.row_dimensions[1].height = 35

    ws.cell(row=2, column=1, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    ws.cell(row=2, column=1).font = Font(italic=True, size=9)
    ws.cell(row=2, column=4, value=f"{len(city_stats)} cities")
    ws.cell(row=2, column=4).font = Font(italic=True, size=9)

    # FUNNEL
    ws.merge_cells('A4:B4')
    c = ws.cell(row=4, column=1, value="AGGREGATE FUNNEL")
    c.font = section_font
    c.fill = section_fill
    c.alignment = center

    funnel_data = [
        ("Hotels Scraped", f"{totals['hotels_scraped']:,}"),
        ("With Website", f"{totals['with_website']:,} ({website_rate}%)"),
        ("Booking Found", f"{totals['booking_found']:,} ({detection_rate}%)"),
    ]
    for i, (label, val) in enumerate(funnel_data):
        r = 5 + i
        ws.cell(row=r, column=1, value=label).font = label_font
        ws.cell(row=r, column=1).border = thin_border
        ws.cell(row=r, column=2, value=val).font = value_font
        ws.cell(row=r, column=2).alignment = right
        ws.cell(row=r, column=2).border = thin_border
        if i == 2:
            ws.cell(row=r, column=1).fill = green_fill
            ws.cell(row=r, column=2).fill = green_fill

    # LEAD QUALITY
    ws.merge_cells('D4:E4')
    c = ws.cell(row=4, column=4, value="LEAD QUALITY")
    c.font = section_font
    c.fill = section_fill
    c.alignment = center

    # Calculate top 25 leads
    top25_leads = sum(c["leads_count"] for c in city_stats if c["city"].lower() in TOP_25_FLORIDA_CITIES)
    top25_booking = sum(c["booking_found"] for c in city_stats if c["city"].lower() in TOP_25_FLORIDA_CITIES)

    quality_data = [
        ("Tier 1 (Known Engine)", f"{totals['tier1']:,} ({tier1_rate}%)"),
        ("Tier 2 (Unknown Engine)", f"{totals['tier2']:,}"),
        ("Total Leads", f"{totals['leads_count']:,}"),
        ("Top 25 Cities Leads", f"{top25_leads:,}"),
    ]
    for i, (label, val) in enumerate(quality_data):
        r = 5 + i
        ws.cell(row=r, column=4, value=label).font = label_font
        ws.cell(row=r, column=4).border = thin_border
        ws.cell(row=r, column=5, value=val).font = value_font
        ws.cell(row=r, column=5).alignment = right
        ws.cell(row=r, column=5).border = thin_border

    # BOOKING ENGINES
    ws.merge_cells('A9:B9')
    c = ws.cell(row=9, column=1, value="BOOKING ENGINES")
    c.font = section_font
    c.fill = section_fill
    c.alignment = center

    # Sort engines by count descending
    sorted_engines = sorted(all_engines.items(), key=lambda x: x[1], reverse=True)
    total_engine_count = sum(all_engines.values())

    for i, (engine, count) in enumerate(sorted_engines[:10]):  # Top 10 engines
        r = 10 + i
        pct = round(count / total_engine_count * 100, 1) if total_engine_count else 0
        engine_display = engine.replace("_", " ").title()
        ws.cell(row=r, column=1, value=engine_display).font = label_font
        ws.cell(row=r, column=1).border = thin_border
        ws.cell(row=r, column=2, value=f"{count:,} ({pct}%)").font = value_font
        ws.cell(row=r, column=2).alignment = right
        ws.cell(row=r, column=2).border = thin_border

    # Calculate row offset for next section
    engine_rows = min(len(sorted_engines), 10)
    next_section_row = 10 + engine_rows + 2

    # CITIES BREAKDOWN
    ws.merge_cells(f'A{next_section_row}:F{next_section_row}')
    c = ws.cell(row=next_section_row, column=1, value="CITIES BREAKDOWN")
    c.font = section_font
    c.fill = section_fill
    c.alignment = center

    # Headers
    headers = ["City", "Top 25", "Leads", "Booking Found", "Tier 1", "Tier 2"]
    ws.column_dimensions["F"].width = 12
    header_row = next_section_row + 1
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=header_row, column=col, value=h)
        c.font = Font(bold=True)
        c.border = thin_border
        c.alignment = center

    # Highlight fill for Top 25 cities
    top25_fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")  # Light yellow

    # City rows - sorted by booking found
    for i, city in enumerate(sorted(city_stats, key=lambda x: x["booking_found"], reverse=True)):
        r = header_row + 1 + i
        city_name_lower = city["city"].lower()
        is_top25 = city_name_lower in TOP_25_FLORIDA_CITIES

        ws.cell(row=r, column=1, value=city["city"]).border = thin_border
        ws.cell(row=r, column=2, value="Yes" if is_top25 else "").border = thin_border
        ws.cell(row=r, column=2).alignment = center
        ws.cell(row=r, column=3, value=city["leads_count"]).border = thin_border
        ws.cell(row=r, column=3).alignment = right
        ws.cell(row=r, column=4, value=city["booking_found"]).border = thin_border
        ws.cell(row=r, column=4).alignment = right
        ws.cell(row=r, column=5, value=city["tier1"]).border = thin_border
        ws.cell(row=r, column=5).alignment = right
        ws.cell(row=r, column=6, value=city["tier2"]).border = thin_border
        ws.cell(row=r, column=6).alignment = right

        # Highlight entire row if Top 25
        if is_top25:
            for col in range(1, 7):
                ws.cell(row=r, column=col).fill = top25_fill

    # Save
    output_path = state_path / f"{state_name} Stats.xlsx"
    wb.save(output_path)

    return output_path


def process_state(state_path: Path):
    """Process a single state folder and generate stats."""
    state_name = state_path.name

    print(f"\n  📊 Processing {state_name}...")

    # Find all city files
    city_files = find_city_files(state_path)

    if not city_files:
        print(f"    No city files found")
        return None

    print(f"    Found {len(city_files)} city files")

    # Extract stats from each
    city_stats = []
    for xlsx in city_files:
        stats = extract_stats_from_excel(xlsx)
        if stats["leads_count"] > 0:
            city_stats.append(stats)

    if not city_stats:
        print(f"    No leads data found")
        return None

    # Generate state stats
    output = generate_state_stats(state_path, city_stats)

    total_leads = sum(c["leads_count"] for c in city_stats)
    total_booking = sum(c["booking_found"] for c in city_stats)
    print(f"    ✅ {state_name} Stats: {len(city_stats)} cities, {total_leads} leads, {total_booking} booking")

    return output


def process_country(country_name: str, state_filter: str = None):
    """Process all states in a country."""
    sadie_path = get_sadie_path()

    # Find country folder (case-insensitive)
    country_path = None
    for item in sadie_path.iterdir():
        if item.is_dir() and item.name.lower() == country_name.lower():
            country_path = item
            break

    if not country_path:
        print(f"❌ Country folder not found: {country_name}")
        return

    print(f"📁 Processing: {country_path}")

    # Find state folders
    for state_dir in sorted(country_path.iterdir()):
        if not state_dir.is_dir():
            continue
        if state_dir.name.startswith("."):
            continue

        # Filter by state if specified
        if state_filter and state_dir.name.lower() != state_filter.lower():
            continue

        process_state(state_dir)


def main():
    parser = argparse.ArgumentParser(description="Generate state-level stats")
    parser.add_argument("--country", default="usa", help="Country to process (default: usa)")
    parser.add_argument("--state", help="Process specific state only (e.g., 'florida')")
    args = parser.parse_args()

    sadie_path = get_sadie_path()
    if not sadie_path.exists():
        print(f"❌ Sadie Lead Gen folder not found: {sadie_path}")
        sys.exit(1)

    process_country(args.country, args.state)

    print("\n✅ State stats complete!")


if __name__ == "__main__":
    main()
