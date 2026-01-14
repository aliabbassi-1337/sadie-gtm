#!/usr/bin/env python3
"""
Load Existing Customers Migration
=================================
Loads existing Sadie customers from Excel file into the database.

Usage:
    uv run python scripts/migrations/load_existing_customers.py
    uv run python scripts/migrations/load_existing_customers.py --file path/to/customers.xlsx
    uv run python scripts/migrations/load_existing_customers.py --dry-run

Requires:
    - openpyxl (pip install openpyxl)
"""

import sys
from pathlib import Path

# Add project root to path for direct script execution
sys.path.insert(0, str(Path(__file__).parent.parent.parent))

import asyncio
import argparse
from typing import Optional, List, Dict
from datetime import datetime

try:
    from openpyxl import load_workbook
except ImportError:
    print("Error: openpyxl not installed. Run: uv add openpyxl")
    sys.exit(1)

from loguru import logger
from db.client import init_db, close_db, get_conn


DEFAULT_FILE = "archive/data/existing_customers.xlsx"


def parse_location(location_str: str) -> tuple:
    """Parse location string like 'Maryland - US' into (state, country)."""
    if not location_str:
        return None, "USA"

    parts = location_str.split(" - ")
    if len(parts) == 2:
        state = parts[0].strip()
        country = "USA" if parts[1].strip() == "US" else parts[1].strip()
        return state, country

    return location_str.strip(), "USA"


def load_customers_from_excel(filepath: str) -> List[Dict]:
    """Load existing customers from Excel file.

    Expected columns (from Sadie customer export):
    - Hotel: customer name
    - Hotel ID: sadie_hotel_id
    - Location: state/country (format: "Maryland - US")
    - Status: "Live", "Churned", etc.
    - Go Live Date: date
    - Latitude: float
    - Longitude: float
    """
    if not Path(filepath).exists():
        logger.error(f"File not found: {filepath}")
        return []

    customers = []

    try:
        wb = load_workbook(filepath, read_only=True, data_only=True)
        ws = wb.active

        # Get headers from first row
        headers = []
        for cell in ws[1]:
            val = str(cell.value).lower().strip() if cell.value else ""
            headers.append(val)

        logger.info(f"Found {len(headers)} columns")

        # Find column indices
        col_map = {}
        for i, h in enumerate(headers):
            col_map[h] = i

        # Required columns
        name_col = col_map.get("hotel")
        if name_col is None:
            logger.error("Could not find 'Hotel' column")
            return []

        # Optional columns
        hotel_id_col = col_map.get("hotel id")
        location_col = col_map.get("location")
        status_col = col_map.get("status")
        go_live_col = col_map.get("go live date")
        lat_col = col_map.get("latitude")
        lon_col = col_map.get("longitude")

        logger.info(f"Column indices: name={name_col}, hotel_id={hotel_id_col}, "
                    f"location={location_col}, status={status_col}, "
                    f"lat={lat_col}, lon={lon_col}")

        # Read data rows
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or not row[name_col]:
                continue

            name = str(row[name_col]).strip()
            if not name:
                continue

            # Parse location (e.g., "Maryland - US" -> state="Maryland")
            state, country = None, "USA"
            if location_col is not None and len(row) > location_col and row[location_col]:
                state, country = parse_location(str(row[location_col]))

            # Map status
            status = "active"
            if status_col is not None and len(row) > status_col and row[status_col]:
                status_val = str(row[status_col]).strip().lower()
                if status_val == "live":
                    status = "active"
                elif status_val == "churned":
                    status = "churned"
                elif status_val == "trial":
                    status = "trial"
                else:
                    status = "active"  # Default unknown statuses to active

            customer = {
                "name": name,
                "sadie_hotel_id": None,
                "latitude": None,
                "longitude": None,
                "state": state,
                "country": country,
                "status": status,
                "go_live_date": None,
            }

            # Extract hotel ID
            if hotel_id_col is not None and len(row) > hotel_id_col and row[hotel_id_col]:
                customer["sadie_hotel_id"] = str(row[hotel_id_col]).strip()

            # Extract coordinates
            if lat_col is not None and len(row) > lat_col and row[lat_col]:
                try:
                    val = row[lat_col]
                    if val != 'x' and val is not None:
                        customer["latitude"] = float(val)
                except (ValueError, TypeError):
                    pass

            if lon_col is not None and len(row) > lon_col and row[lon_col]:
                try:
                    val = row[lon_col]
                    if val != 'x' and val is not None:
                        customer["longitude"] = float(val)
                except (ValueError, TypeError):
                    pass

            # Extract go-live date
            if go_live_col is not None and len(row) > go_live_col and row[go_live_col]:
                val = row[go_live_col]
                if isinstance(val, datetime):
                    customer["go_live_date"] = val.date()

            customers.append(customer)

        wb.close()

    except Exception as e:
        logger.error(f"Error reading Excel file: {e}")
        import traceback
        traceback.print_exc()
        return []

    logger.info(f"Loaded {len(customers)} customers from {filepath}")

    # Stats
    with_coords = sum(1 for c in customers if c["latitude"] and c["longitude"])
    active = sum(1 for c in customers if c["status"] == "active")
    logger.info(f"  {with_coords} have coordinates")
    logger.info(f"  {active} are active")

    return customers


async def insert_customers(customers: List[Dict], dry_run: bool = False) -> int:
    """Insert customers into database."""
    if not customers:
        return 0

    if dry_run:
        logger.info("DRY RUN - would insert the following customers:")
        for c in customers[:10]:
            coord_str = f"@ {c['latitude']}, {c['longitude']}" if c['latitude'] else "(no coords)"
            logger.info(f"  - {c['name']} ({c['state']}) {coord_str} [{c['status']}]")
        if len(customers) > 10:
            logger.info(f"  ... and {len(customers) - 10} more")
        return len(customers)

    inserted = 0
    skipped = 0

    async with get_conn() as conn:
        for customer in customers:
            try:
                lat = customer["latitude"]
                lon = customer["longitude"]

                # Check if already exists
                existing = await conn.fetchval(
                    "SELECT id FROM existing_customers WHERE name = $1",
                    customer["name"]
                )
                if existing:
                    skipped += 1
                    continue

                # Insert with PostGIS point
                if lat and lon:
                    await conn.execute("""
                        INSERT INTO existing_customers (
                            name, sadie_hotel_id, location, city, state, country, status, go_live_date
                        ) VALUES (
                            $1, $2, ST_Point($3, $4)::geography, NULL, $5, $6, $7, $8
                        )
                    """,
                        customer["name"],
                        customer["sadie_hotel_id"],
                        lon,  # longitude first for ST_Point
                        lat,
                        customer["state"],
                        customer["country"],
                        customer["status"],
                        customer["go_live_date"],
                    )
                else:
                    await conn.execute("""
                        INSERT INTO existing_customers (
                            name, sadie_hotel_id, location, city, state, country, status, go_live_date
                        ) VALUES (
                            $1, $2, NULL, NULL, $3, $4, $5, $6
                        )
                    """,
                        customer["name"],
                        customer["sadie_hotel_id"],
                        customer["state"],
                        customer["country"],
                        customer["status"],
                        customer["go_live_date"],
                    )
                inserted += 1

            except Exception as e:
                logger.error(f"Error inserting {customer['name']}: {e}")

    if skipped > 0:
        logger.info(f"Skipped {skipped} customers (already exist)")

    return inserted


async def run(filepath: str, dry_run: bool = False):
    """Load customers from Excel and insert into database."""
    logger.info(f"Loading existing customers from: {filepath}")

    # Load from Excel
    customers = load_customers_from_excel(filepath)

    if not customers:
        logger.error("No customers loaded from file")
        return

    # Initialize database
    await init_db()

    try:
        # Check current count
        async with get_conn() as conn:
            result = await conn.fetchval("SELECT COUNT(*) FROM existing_customers")
            logger.info(f"Current customers in database: {result}")

        # Insert customers
        inserted = await insert_customers(customers, dry_run=dry_run)

        if not dry_run:
            # Verify
            async with get_conn() as conn:
                total = await conn.fetchval("SELECT COUNT(*) FROM existing_customers")
                with_location = await conn.fetchval(
                    "SELECT COUNT(*) FROM existing_customers WHERE location IS NOT NULL"
                )
                logger.info(f"Customers in database after insert: {total}")
                logger.info(f"Customers with location: {with_location}")

        logger.info("=" * 60)
        logger.info("MIGRATION COMPLETE" if not dry_run else "DRY RUN COMPLETE")
        logger.info("=" * 60)
        logger.info(f"Customers in file: {len(customers)}")
        logger.info(f"Customers inserted: {inserted}")
        logger.info("=" * 60)

    finally:
        await close_db()


def main():
    parser = argparse.ArgumentParser(description="Load existing customers from Excel into database")
    parser.add_argument(
        "--file", "-f",
        default=DEFAULT_FILE,
        help=f"Path to Excel file (default: {DEFAULT_FILE})"
    )
    parser.add_argument(
        "--dry-run", "-n",
        action="store_true",
        help="Show what would be inserted without actually inserting"
    )

    args = parser.parse_args()

    asyncio.run(run(filepath=args.file, dry_run=args.dry_run))


if __name__ == "__main__":
    main()
