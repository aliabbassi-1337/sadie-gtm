"""Reporting service for generating and uploading Excel reports."""

import os
import tempfile
from abc import ABC, abstractmethod
from typing import List, Optional

from loguru import logger
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from services.reporting import repo
from db.models.reporting import HotelLead, CityStats, EngineCount, ReportStats, LaunchableHotel, EnrichmentStats
from infra.s3 import upload_file
from infra import slack


class IService(ABC):
    """Reporting Service - Generate and deliver reports to stakeholders."""

    @abstractmethod
    async def export_city(self, city: str, state: str, country: str = "United States") -> tuple[str, int]:
        """Generate Excel report for a city and upload to S3.

        Returns:
            Tuple of (s3_uri, lead_count)
        """
        pass

    @abstractmethod
    async def export_state(self, state: str, country: str = "United States", source_pattern: str = None) -> tuple[str, int]:
        """Generate Excel report for an entire state and upload to S3.

        Returns:
            Tuple of (s3_uri, lead_count)
        """
        pass

    @abstractmethod
    async def export_by_booking_engine(
        self,
        booking_engine: str,
        source_pattern: str = "%crawl%",
    ) -> tuple[str, int]:
        """Generate Excel report for hotels with a specific booking engine from crawl data.

        Creates files like: cloudbeds_crawldata.xlsx

        Returns:
            Tuple of (s3_uri, lead_count)
        """
        pass

    @abstractmethod
    async def export_by_source(
        self,
        source_pattern: str,
        filename: str = None,
    ) -> tuple[str, int]:
        """Generate Excel report for hotels from a specific source.

        Creates files like: ipms247_leads.xlsx

        Returns:
            Tuple of (s3_uri, lead_count)
        """
        pass

    @abstractmethod
    async def export_all_states(
        self,
        country: str = "United States",
        source_pattern: str = None,
    ) -> dict:
        """Export all states that have hotels.

        Args:
            country: Country name (e.g. "United States", "Australia")
            source_pattern: Filter by source pattern

        Returns:
            Dict with 'states' count and 'total_leads' count
        """
        pass

    @abstractmethod
    def send_slack_notification(
        self,
        location: str,
        lead_count: int,
        s3_uri: str,
        channel: str = "#leads",
    ) -> bool:
        """Send export notification to Slack.

        Returns True if sent successfully.
        """
        pass

    # =========================================================================
    # LAUNCHER METHODS
    # =========================================================================

    @abstractmethod
    async def get_launchable_hotels(self, limit: int = 100) -> List[LaunchableHotel]:
        """Get hotels ready to be launched (fully enriched with all data).

        A hotel is launchable when it has:
        - status = 0 (pending)
        - A record in hotel_booking_engines
        - A record in hotel_room_count (with status=1)
        - A record in hotel_customer_proximity
        """
        pass

    @abstractmethod
    async def get_launchable_count(self) -> int:
        """Count hotels ready to be launched."""
        pass

    @abstractmethod
    async def launch_hotels(self, hotel_ids: List[int]) -> int:
        """Atomically claim and launch specific hotels (multi-worker safe).

        Uses FOR UPDATE SKIP LOCKED so multiple EC2 instances can run concurrently.
        Returns the number of hotels actually launched.
        """
        pass

    @abstractmethod
    async def launch_ready(self, limit: int = 100) -> int:
        """Atomically claim and launch ready hotels (multi-worker safe).

        Uses FOR UPDATE SKIP LOCKED so multiple EC2 instances can run concurrently.
        Returns the number of hotels launched.
        """
        pass

    @abstractmethod
    async def get_launched_count(self) -> int:
        """Count hotels that have been launched."""
        pass

    # =========================================================================
    # PIPELINE STATUS METHODS
    # =========================================================================

    @abstractmethod
    async def get_pipeline_summary(self) -> list:
        """Get count of hotels at each pipeline stage."""
        pass

    @abstractmethod
    async def get_pipeline_by_source(self) -> list:
        """Get pipeline breakdown by source."""
        pass

    @abstractmethod
    async def get_pipeline_by_source_name(self, source: str) -> list:
        """Get pipeline breakdown for a specific source."""
        pass

    # =========================================================================
    # ENRICHMENT STATS METHODS
    # =========================================================================

    @abstractmethod
    async def get_enrichment_stats(self, source_pattern: str = None) -> List[EnrichmentStats]:
        """Get enrichment stats by booking engine.

        Args:
            source_pattern: Optional source pattern filter (e.g., '%crawl%')

        Returns:
            List of EnrichmentStats, one per booking engine
        """
        pass

    @abstractmethod
    async def export_enrichment_stats(self, source_pattern: str = None) -> tuple[str, int]:
        """Export enrichment stats to Excel and upload to S3.

        Returns:
            Tuple of (s3_uri, engine_count)
        """
        pass


class Service(IService):
    """Implementation of the reporting service."""

    def __init__(self) -> None:
        pass

    async def export_city(self, city: str, state: str, country: str = "United States") -> tuple[str, int]:
        """Generate Excel report for a city and upload to S3.

        Returns:
            Tuple of (s3_uri, lead_count)
        """
        logger.info(f"Generating report for {city}, {state}")

        # Get data from database
        leads = await repo.get_leads_for_city(city, state)
        stats = await repo.get_city_stats(city, state)
        top_engines = await repo.get_top_engines_for_city(city, state)

        report_stats = ReportStats(
            location_name=city,
            stats=stats,
            top_engines=top_engines,
        )

        # Generate Excel workbook
        workbook = self._create_workbook(leads, report_stats)

        # Save to temp file and upload
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            workbook.save(tmp.name)
            tmp_path = tmp.name

        try:
            # S3 path: HotelLeadGen/{country}/{state}/{city}.xlsx
            s3_key = f"HotelLeadGen/{country}/{state}/{city}.xlsx"
            s3_uri = upload_file(tmp_path, s3_key)
            logger.info(f"Uploaded city report to {s3_uri}")
            return s3_uri, len(leads)
        finally:
            os.unlink(tmp_path)

    async def export_state(self, state: str, country: str = "United States", source_pattern: str = None) -> tuple[str, int]:
        """Generate Excel report for an entire state and upload to S3.

        Uses s5cmd for fast upload with fallback to boto3.

        Returns:
            Tuple of (s3_uri, lead_count)
        """
        import subprocess

        logger.info(f"Generating state report for {state}")

        # Get data from database
        leads = await repo.get_leads_for_state(state, source_pattern=source_pattern, country=country)
        stats = await repo.get_state_stats(state, source_pattern=source_pattern, country=country)
        top_engines = await repo.get_top_engines_for_state(state, source_pattern=source_pattern, country=country)
        if source_pattern:
            funnel = await repo.get_detection_funnel_by_source(state, source_pattern, country=country)
        else:
            funnel = await repo.get_detection_funnel(state, country=country)

        logger.info(f"Found {len(leads)} leads for {state}")

        report_stats = ReportStats(
            location_name=state,
            stats=stats,
            top_engines=top_engines,
            funnel=funnel,
        )

        # Generate Excel workbook
        workbook = self._create_workbook(leads, report_stats)

        # Save to temp file and upload
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            workbook.save(tmp.name)
            tmp_path = tmp.name

        try:
            if source_pattern:
                source_name = source_pattern.replace('%', '').replace('_', '-')
                filename = f"{state}_{source_name}.xlsx"
            else:
                filename = f"{state}.xlsx"
            s3_uri = f"s3://sadie-gtm/HotelLeadGen/{country}/{state}/{filename}"

            # Try s5cmd first (faster)
            result = subprocess.run(
                ["s5cmd", "cp", tmp_path, s3_uri],
                capture_output=True,
                text=True,
            )

            if result.returncode != 0:
                # Fallback to boto3
                logger.warning(f"s5cmd failed, using boto3: {result.stderr}")
                s3_key = f"HotelLeadGen/{country}/{state}/{state}.xlsx"
                s3_uri = upload_file(tmp_path, s3_key)

            logger.info(f"Uploaded state report to {s3_uri}")
            return s3_uri, len(leads)
        finally:
            os.unlink(tmp_path)

    async def export_country_leads(self, country: str, db_country: str) -> tuple[str, int]:
        """Generate combined Excel report for ALL leads in a country.

        Creates files like: USA/USA_leads.xlsx, Australia/Australia_leads.xlsx
        """
        import subprocess

        logger.info(f"Generating country-level report for {country}")

        leads = await repo.get_leads_for_country(db_country)
        logger.info(f"Found {len(leads)} total leads for {country}")

        if not leads:
            return "", 0

        workbook = self._create_crawl_workbook(leads, f"{country} All Leads")

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            workbook.save(tmp.name)
            tmp_path = tmp.name

        try:
            filename = f"{country}_leads.xlsx"
            s3_uri = f"s3://sadie-gtm/HotelLeadGen/{country}/{filename}"

            result = subprocess.run(
                ["s5cmd", "cp", tmp_path, s3_uri],
                capture_output=True,
                text=True,
            )

            if result.returncode != 0:
                logger.warning(f"s5cmd failed, using boto3: {result.stderr}")
                s3_key = f"HotelLeadGen/{country}/{filename}"
                s3_uri = upload_file(tmp_path, s3_key)

            logger.info(f"Uploaded country report to {s3_uri}")
            return s3_uri, len(leads)
        finally:
            os.unlink(tmp_path)

    async def export_by_booking_engine(
        self,
        booking_engine: str,
        source_pattern: str = "%crawl%",
    ) -> tuple[str, int]:
        """Generate Excel report for hotels with a specific booking engine from crawl data.

        Creates files like: cloudbeds_leads.xlsx
        """
        import subprocess

        # Normalize engine name for query (e.g., "cloudbeds" -> "Cloudbeds")
        engine_name = booking_engine.title()
        if booking_engine.lower() == "rms":
            engine_name = "RMS Cloud"
        elif booking_engine.lower() == "siteminder":
            engine_name = "SiteMinder"

        logger.info(f"Generating crawl data report for {engine_name}")

        # Get leads
        leads = await repo.get_leads_by_booking_engine(engine_name, source_pattern)
        logger.info(f"Found {len(leads)} leads for {engine_name}")

        if not leads:
            logger.warning(f"No leads found for {engine_name} with source pattern {source_pattern}")
            return "", 0

        # Create workbook with simplified leads sheet (no stats for crawl data)
        workbook = self._create_crawl_workbook(leads, engine_name)

        # Save to temp file and upload
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            workbook.save(tmp.name)
            tmp_path = tmp.name

        try:
            # Filename: cloudbeds_leads.xlsx
            filename = f"{booking_engine.lower()}_leads.xlsx"
            s3_uri = f"s3://sadie-gtm/HotelLeadGen/booking-engines/{filename}"

            # Try s5cmd first (faster)
            result = subprocess.run(
                ["s5cmd", "cp", tmp_path, s3_uri],
                capture_output=True,
                text=True,
            )

            if result.returncode != 0:
                logger.warning(f"s5cmd failed, using boto3: {result.stderr}")
                s3_key = f"HotelLeadGen/booking-engines/{filename}"
                s3_uri = upload_file(tmp_path, s3_key)

            logger.info(f"Uploaded crawl data report to {s3_uri}")
            return s3_uri, len(leads)
        finally:
            os.unlink(tmp_path)

    async def export_by_source(
        self,
        source_pattern: str,
        filename: str = None,
    ) -> tuple[str, int]:
        """Generate Excel report for hotels from a specific source.

        Creates files like: ipms247_leads.xlsx
        """
        import subprocess

        logger.info(f"Generating source report for {source_pattern}")

        # Get leads
        leads = await repo.get_leads_by_source(source_pattern)
        logger.info(f"Found {len(leads)} leads for source {source_pattern}")

        if not leads:
            logger.warning(f"No leads found for source {source_pattern}")
            return "", 0

        # Determine source name for filename
        source_name = source_pattern.replace('%', '').replace('_', '')
        if not filename:
            filename = f"{source_name}_leads.xlsx"

        # Create workbook
        workbook = self._create_crawl_workbook(leads, source_name)

        # Save to temp file and upload
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            workbook.save(tmp.name)
            tmp_path = tmp.name

        try:
            s3_uri = f"s3://sadie-gtm/HotelLeadGen/booking-engines/{filename}"

            result = subprocess.run(
                ["s5cmd", "cp", tmp_path, s3_uri],
                capture_output=True,
                text=True,
            )

            if result.returncode != 0:
                logger.warning(f"s5cmd failed, using boto3: {result.stderr}")
                s3_key = f"HotelLeadGen/booking-engines/{filename}"
                s3_uri = upload_file(tmp_path, s3_key)

            logger.info(f"Uploaded source report to {s3_uri}")
            return s3_uri, len(leads)
        finally:
            os.unlink(tmp_path)

    async def export_all_states(
        self,
        country: str = "United States",
        source_pattern: str = None,
    ) -> dict:
        """Export all states that have hotels.

        Args:
            country: Country name (e.g. "United States", "Australia")
            source_pattern: Filter by source pattern

        Returns:
            Dict with 'states' count and 'total_leads' count
        """
        states = await repo.get_distinct_states(country=country)

        logger.info(f"Found {len(states)} states with hotels")

        total_leads = 0
        exports = []

        for state in states:
            if not state:
                continue
            try:
                logger.info(f"Exporting {state}...")
                s3_uri, lead_count = await self.export_state(state, country, source_pattern=source_pattern)
                if lead_count > 0:
                    exports.append((state, s3_uri, lead_count))
                    total_leads += lead_count
                    logger.info(f"  {state}: {lead_count} leads -> {s3_uri}")
                else:
                    logger.info(f"  {state}: no leads")
            except Exception as e:
                logger.error(f"  {state}: failed - {e}")

        # Summary
        logger.info("\n" + "=" * 50)
        logger.info("EXPORT SUMMARY")
        logger.info("=" * 50)
        for state, s3_uri, count in exports:
            logger.info(f"  {state}: {count} leads")
        logger.info(f"Total: {total_leads} leads across {len(exports)} states")

        return {"states": len(exports), "total_leads": total_leads}

    def _create_crawl_workbook(self, leads: List[HotelLead], engine_name: str) -> Workbook:
        """Create Excel workbook for crawl data (simpler format, no stats)."""
        workbook = Workbook()

        # Create leads sheet
        leads_sheet = workbook.active
        leads_sheet.title = "Leads"
        self._populate_crawl_leads_sheet(leads_sheet, leads)

        return workbook

    def _populate_crawl_leads_sheet(self, sheet, leads: List[HotelLead]) -> None:
        """Populate the crawl leads sheet with hotel data."""
        # Define all possible columns with their extractors
        # Category moved to the right (after Booking Engine)
        all_columns = [
            ("Hotel", lambda l: l.hotel_name),
            ("Address", lambda l: l.address or ""),
            ("City", lambda l: l.city or ""),
            ("State", lambda l: l.state or ""),
            ("Country", lambda l: l.country or ""),
            ("Phone", lambda l: l.phone_website or l.phone_google or ""),
            ("Email", lambda l: l.email or ""),
            ("Website", lambda l: l.website or ""),
            ("Rating", lambda l: l.rating or ""),
            ("Reviews", lambda l: l.review_count or ""),
            ("Room Count", lambda l: l.room_count or ""),
            ("Room Count Source", lambda l: l.room_count_source or ""),
            ("Room Count Confidence", lambda l: l.room_count_confidence or ""),
            ("Booking URL", lambda l: l.booking_url or ""),
            ("Booking Engine", lambda l: l.booking_engine_name or ""),
            ("Category", lambda l: l.category or ""),
            ("Active", lambda l: "Yes" if l.is_active is True else ("No" if l.is_active is False else "")),
            ("Has Availability", lambda l: "Yes" if l.has_availability is True else ("No" if l.has_availability is False else "")),
        ]

        # Filter out columns where ALL values are empty
        columns = []
        for header, extractor in all_columns:
            has_data = any(extractor(lead) for lead in leads)
            if has_data:
                columns.append((header, extractor))

        headers = [c[0] for c in columns]
        extractors = [c[1] for c in columns]

        # Style definitions
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Write headers
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # Write data rows
        import re
        _illegal_xml_re = re.compile(
            r"[\x00-\x08\x0b\x0c\x0e-\x1f\x7f-\x9f]"
        )
        for row, lead in enumerate(leads, 2):
            for col, extractor in enumerate(extractors, 1):
                value = extractor(lead)
                if isinstance(value, str):
                    value = _illegal_xml_re.sub("", value)
                cell = sheet.cell(row=row, column=col, value=value)
                cell.border = thin_border

        # Auto-adjust column widths
        for col in range(1, len(headers) + 1):
            max_length = len(headers[col - 1])
            for row in range(2, min(len(leads) + 2, 100)):  # Sample first 100 rows
                cell_value = sheet.cell(row=row, column=col).value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))
            sheet.column_dimensions[get_column_letter(col)].width = min(max_length + 2, 60)

    def send_slack_notification(
        self,
        location: str,
        lead_count: int,
        s3_uri: str,
        channel: str = "#leads",
    ) -> bool:
        """Send export notification to Slack."""
        message = (
            f"*Lead Export Complete*\n"
            f"• Location: {location}\n"
            f"• Leads: {lead_count}\n"
            f"• File: `{s3_uri}`"
        )
        return slack.send_message(message, channel)

    def _create_workbook(
        self,
        leads: List[HotelLead],
        report_stats: ReportStats,
    ) -> Workbook:
        """Create Excel workbook with leads and stats tabs."""
        workbook = Workbook()

        # Create leads sheet (first tab)
        leads_sheet = workbook.active
        leads_sheet.title = "Leads"
        self._populate_leads_sheet(leads_sheet, leads)

        # Create stats sheet (second tab)
        stats_sheet = workbook.create_sheet(title="Stats")
        self._populate_stats_sheet(stats_sheet, report_stats)

        return workbook

    def _populate_leads_sheet(self, sheet, leads: List[HotelLead]) -> None:
        """Populate the leads sheet with hotel data."""
        import re
        _illegal_xml_re = re.compile(
            r"[\x00-\x08\x0b\x0c\x0e-\x1f\x7f-\x9f]"
        )

        # Define all columns with extractors — skip columns where ALL values are blank
        all_columns = [
            ("Hotel", lambda l: l.hotel_name),
            ("Address", lambda l: l.address or ""),
            ("City", lambda l: l.city or ""),
            ("State", lambda l: l.state or ""),
            ("Country", lambda l: l.country or ""),
            ("Phone", lambda l: l.phone_website or l.phone_google or ""),
            ("Email", lambda l: l.email or ""),
            ("Website", lambda l: l.website or ""),
            ("Rating", lambda l: l.rating or ""),
            ("Reviews", lambda l: l.review_count or ""),
            ("Room Count", lambda l: l.room_count or ""),
            ("Room Count Source", lambda l: l.room_count_source or ""),
            ("Room Count Confidence", lambda l: l.room_count_confidence or ""),
            ("Booking URL", lambda l: l.booking_url or ""),
            ("Booking Engine", lambda l: l.booking_engine_name or ""),
            ("Category", lambda l: l.category or ""),
            ("Proximity", lambda l: self._format_proximity(l)),
            ("Active", lambda l: "Yes" if l.is_active is True else ("No" if l.is_active is False else "")),
            ("Has Availability", lambda l: "Yes" if l.has_availability is True else ("No" if l.has_availability is False else "")),
        ]

        # Filter out columns where ALL values are empty
        columns = []
        for header, extractor in all_columns:
            has_data = any(extractor(lead) for lead in leads)
            if has_data:
                columns.append((header, extractor))

        headers = [c[0] for c in columns]
        extractors = [c[1] for c in columns]

        # Style definitions
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Write headers
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # Write data rows
        for row, lead in enumerate(leads, 2):
            for col, extractor in enumerate(extractors, 1):
                value = extractor(lead)
                if isinstance(value, str):
                    value = _illegal_xml_re.sub("", value)
                cell = sheet.cell(row=row, column=col, value=value)
                cell.border = thin_border

        # Auto-adjust column widths
        for col in range(1, len(headers) + 1):
            max_length = len(headers[col - 1])
            for row in range(2, min(len(leads) + 2, 100)):  # Sample first 100 rows
                cell_value = sheet.cell(row=row, column=col).value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))
            sheet.column_dimensions[get_column_letter(col)].width = min(max_length + 2, 60)

    def _format_proximity(self, lead: HotelLead) -> str:
        """Format proximity string for a lead."""
        if not lead.nearest_customer_name or lead.nearest_customer_distance_km is None:
            return "No nearby customer"

        distance = float(lead.nearest_customer_distance_km)
        return f"Nearest: {lead.nearest_customer_name} ({distance:.1f}km)"

    def _populate_stats_sheet(self, sheet, report_stats: ReportStats) -> None:
        """Populate the stats sheet with detection funnel dashboard."""
        location = report_stats.location_name.upper()
        funnel = report_stats.funnel or {}
        stats = report_stats.stats

        # Style definitions
        title_font = Font(bold=True, size=14)
        section_font = Font(bold=True, size=11)

        # Title
        sheet.cell(row=1, column=1, value=f"LEAD GENERATION DASHBOARD - {location}")
        sheet.cell(row=1, column=1).font = title_font
        sheet.merge_cells("A1:F1")

        row = 3

        # DETECTION FUNNEL Section
        sheet.cell(row=row, column=1, value="DETECTION FUNNEL")
        sheet.cell(row=row, column=1).font = section_font
        row += 1

        total = funnel.get("total_hotels", 0)
        with_website = funnel.get("with_website", 0)
        detection_attempted = funnel.get("detection_attempted", 0)
        engine_found = funnel.get("engine_found", 0)
        ota_found = funnel.get("ota_found", 0)
        no_engine_found = funnel.get("no_engine_found", 0)
        pending = funnel.get("pending_detection", 0)
        launched = funnel.get("launched", 0)

        # Funnel rows
        sheet.cell(row=row, column=1, value="Total Hotels")
        sheet.cell(row=row, column=2, value=total)
        sheet.cell(row=row, column=3, value="100%")
        row += 1

        sheet.cell(row=row, column=1, value="  → With Website")
        sheet.cell(row=row, column=2, value=with_website)
        sheet.cell(row=row, column=3, value=f"{with_website/total*100:.1f}%" if total else "0%")
        row += 1

        sheet.cell(row=row, column=1, value="    → Detection Attempted")
        sheet.cell(row=row, column=2, value=detection_attempted)
        sheet.cell(row=row, column=3, value=f"{detection_attempted/with_website*100:.1f}%" if with_website else "0%")
        row += 1

        sheet.cell(row=row, column=1, value="      ✓ Engine Found")
        sheet.cell(row=row, column=2, value=engine_found)
        sheet.cell(row=row, column=3, value=f"{engine_found/detection_attempted*100:.1f}%" if detection_attempted else "0%")
        row += 1

        sheet.cell(row=row, column=1, value="      ✗ No Engine Found")
        sheet.cell(row=row, column=2, value=no_engine_found)
        sheet.cell(row=row, column=3, value=f"{no_engine_found/detection_attempted*100:.1f}%" if detection_attempted else "0%")
        row += 1

        sheet.cell(row=row, column=1, value="    → Pending Detection")
        sheet.cell(row=row, column=2, value=pending)
        sheet.cell(row=row, column=3, value="(not attempted)")
        row += 1

        sheet.cell(row=row, column=1, value="        → Launched")
        sheet.cell(row=row, column=2, value=launched)
        row += 2

        # FAILURE BREAKDOWN Section
        sheet.cell(row=row, column=1, value="FAILURE BREAKDOWN")
        sheet.cell(row=row, column=1).font = section_font
        row += 1

        total_failures = (funnel.get("http_403", 0) + funnel.get("http_429", 0) +
                         funnel.get("junk_url", 0) + funnel.get("junk_domain", 0) +
                         funnel.get("non_hotel_name", 0) + funnel.get("timeout_err", 0) +
                         funnel.get("server_5xx", 0) + funnel.get("browser_err", 0))

        failures = [
            ("HTTP 403 (Bot Protection)", funnel.get("http_403", 0)),
            ("HTTP 429 (Rate Limited)", funnel.get("http_429", 0)),
            ("Junk Booking URL", funnel.get("junk_url", 0)),
            ("Junk Domain", funnel.get("junk_domain", 0)),
            ("Non-Hotel Name", funnel.get("non_hotel_name", 0)),
            ("Timeout", funnel.get("timeout_err", 0)),
            ("Server Error (5xx)", funnel.get("server_5xx", 0)),
            ("Browser Exception", funnel.get("browser_err", 0)),
        ]

        for label, count in failures:
            pct = f"({count/total_failures*100:.1f}%)" if total_failures else ""
            sheet.cell(row=row, column=1, value=label)
            sheet.cell(row=row, column=2, value=count)
            sheet.cell(row=row, column=3, value=pct)
            row += 1

        row += 1

        # 70% TARGET ANALYSIS Section
        sheet.cell(row=row, column=1, value="70% TARGET ANALYSIS")
        sheet.cell(row=row, column=1).font = section_font
        row += 1

        target_70 = int(with_website * 0.7)
        gap = target_70 - engine_found
        available_pool = pending + no_engine_found

        sheet.cell(row=row, column=1, value="Target (70% of with-website)")
        sheet.cell(row=row, column=2, value=target_70)
        row += 1

        sheet.cell(row=row, column=1, value="Current Engine Found")
        sheet.cell(row=row, column=2, value=engine_found)
        row += 1

        sheet.cell(row=row, column=1, value="Gap to 70%")
        sheet.cell(row=row, column=2, value=gap)
        row += 1

        sheet.cell(row=row, column=1, value="Available Pool (pending+failed)")
        sheet.cell(row=row, column=2, value=available_pool)
        row += 2

        # LEAD QUALITY Section
        sheet.cell(row=row, column=1, value="LEAD QUALITY")
        sheet.cell(row=row, column=1).font = section_font
        row += 1

        total_with_booking = stats.tier_1_count + stats.tier_2_count
        tier_1_pct = stats.tier_1_count / total_with_booking * 100 if total_with_booking else 0
        tier_2_pct = stats.tier_2_count / total_with_booking * 100 if total_with_booking else 0

        sheet.cell(row=row, column=1, value="Tier 1 (Known Engine)")
        sheet.cell(row=row, column=2, value=stats.tier_1_count)
        sheet.cell(row=row, column=3, value=f"{tier_1_pct:.1f}%")
        row += 1

        sheet.cell(row=row, column=1, value="Tier 2 (Unknown Engine)")
        sheet.cell(row=row, column=2, value=stats.tier_2_count)
        sheet.cell(row=row, column=3, value=f"{tier_2_pct:.1f}%")
        row += 2

        # TOP ENGINES Section
        sheet.cell(row=row, column=1, value="TOP ENGINES")
        sheet.cell(row=row, column=1).font = section_font
        row += 1

        for engine in report_stats.top_engines:
            sheet.cell(row=row, column=1, value=engine.engine_name)
            sheet.cell(row=row, column=2, value=engine.hotel_count)
            row += 1

        # Auto-adjust column widths
        sheet.column_dimensions["A"].width = 35
        sheet.column_dimensions["B"].width = 12
        sheet.column_dimensions["C"].width = 15

    # =========================================================================
    # LAUNCHER METHODS
    # =========================================================================

    async def get_launchable_hotels(self, limit: int = 100) -> List[LaunchableHotel]:
        """Get hotels ready to be launched (fully enriched with all data)."""
        return await repo.get_launchable_hotels(limit=limit)

    async def get_launchable_count(self) -> int:
        """Count hotels ready to be launched."""
        return await repo.get_launchable_count()

    async def launch_hotels(self, hotel_ids: List[int]) -> int:
        """Atomically claim and launch specific hotels (multi-worker safe).

        Uses FOR UPDATE SKIP LOCKED so multiple EC2 instances can run concurrently.
        Returns the number of hotels actually launched.
        """
        if not hotel_ids:
            return 0

        launched_ids = await repo.launch_hotels(hotel_ids)
        if launched_ids:
            logger.info(f"Launched {len(launched_ids)} hotels: {launched_ids}")
        return len(launched_ids)

    async def launch_ready(self, limit: int = 100) -> int:
        """Atomically claim and launch ready hotels (multi-worker safe).

        Uses FOR UPDATE SKIP LOCKED so multiple EC2 instances can run concurrently.
        Returns the number of hotels launched.
        """
        launched_ids = await repo.launch_ready_hotels(limit=limit)
        if launched_ids:
            logger.info(f"Launched {len(launched_ids)} hotels: {launched_ids}")
        else:
            logger.info("No hotels ready to launch")
        return len(launched_ids)

    async def get_launched_count(self) -> int:
        """Count hotels that have been launched."""
        return await repo.get_launched_count()

    async def get_takedown_count(self) -> int:
        """Count launched hotels without an active booking engine."""
        return await repo.get_takedown_count()

    async def get_takedown_candidates(self, limit: int = 100) -> list:
        """Get launched hotels without an active booking engine."""
        return await repo.get_takedown_candidates(limit=limit)

    async def takedown_hotels_without_engine(self, limit: int = 10000) -> int:
        """Take down launched hotels without an active booking engine.

        Sets status back to 0 (pending). Returns count of taken-down hotels.
        """
        taken_down_ids = await repo.takedown_hotels_without_engine(limit=limit)
        if taken_down_ids:
            logger.info(f"Taken down {len(taken_down_ids)} hotels without active engine")
        else:
            logger.info("No hotels to take down")
        return len(taken_down_ids)

    # =========================================================================
    # PIPELINE STATUS METHODS
    # =========================================================================

    async def get_pipeline_summary(self) -> list:
        """Get count of hotels at each pipeline stage."""
        return await repo.get_pipeline_summary()

    async def get_pipeline_by_source(self) -> list:
        """Get pipeline breakdown by source."""
        return await repo.get_pipeline_by_source()

    async def get_pipeline_by_source_name(self, source: str) -> list:
        """Get pipeline breakdown for a specific source."""
        return await repo.get_pipeline_by_source_name(source)

    # =========================================================================
    # ENRICHMENT STATS METHODS
    # =========================================================================

    async def get_enrichment_stats(self, source_pattern: str = None) -> List[EnrichmentStats]:
        """Get enrichment stats by booking engine."""
        return await repo.get_enrichment_stats_by_engine(source_pattern=source_pattern)

    async def export_enrichment_stats(self, source_pattern: str = None) -> tuple[str, int]:
        """Export enrichment stats to Excel and upload to S3."""
        import subprocess

        logger.info("Generating enrichment stats report")

        stats = await self.get_enrichment_stats(source_pattern=source_pattern)
        if not stats:
            logger.warning("No enrichment stats found")
            return "", 0

        logger.info(f"Found stats for {len(stats)} booking engines")

        workbook = self._create_enrichment_stats_workbook(stats)

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            workbook.save(tmp.name)
            tmp_path = tmp.name

        try:
            if source_pattern:
                source_tag = source_pattern.replace('%', '').replace('_', '')
                filename = f"enrichment_stats_{source_tag}.xlsx"
            else:
                filename = "enrichment_stats.xlsx"
            s3_uri = f"s3://sadie-gtm/HotelLeadGen/reports/{filename}"

            result = subprocess.run(
                ["s5cmd", "cp", tmp_path, s3_uri],
                capture_output=True,
                text=True,
            )

            if result.returncode != 0:
                logger.warning(f"s5cmd failed, using boto3: {result.stderr}")
                s3_key = f"HotelLeadGen/reports/{filename}"
                s3_uri = upload_file(tmp_path, s3_key)

            logger.info(f"Uploaded enrichment stats report to {s3_uri}")
            return s3_uri, len(stats)
        finally:
            os.unlink(tmp_path)

    def _create_enrichment_stats_workbook(self, stats: List[EnrichmentStats]) -> Workbook:
        """Create Excel workbook for enrichment stats."""
        workbook = Workbook()

        sheet = workbook.active
        sheet.title = "Enrichment Stats"
        self._populate_enrichment_stats_sheet(sheet, stats)

        return workbook

    def _populate_enrichment_stats_sheet(self, sheet, stats: List[EnrichmentStats]) -> None:
        """Populate the enrichment stats sheet."""
        headers = [
            "Engine", "Total", "Live", "Pending", "Error",
            "Has Name", "Has Email", "Has Phone", "Has Contact",
            "Has City", "Has State", "Has Country",
            "Has Website", "Has Address", "Has Coordinates",
            "Has Booking Engine", "Has Room Count"
        ]

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        for row, stat in enumerate(stats, 2):
            data = [
                stat.engine_name,
                stat.total_hotels,
                stat.live,
                stat.pending,
                stat.error,
                stat.has_name,
                stat.has_email,
                stat.has_phone,
                stat.has_contact,
                stat.has_city,
                stat.has_state,
                stat.has_country,
                stat.has_website,
                stat.has_address,
                stat.has_coordinates,
                stat.has_booking_engine,
                stat.has_room_count,
            ]
            for col, value in enumerate(data, 1):
                cell = sheet.cell(row=row, column=col, value=value)
                cell.border = thin_border
                if col > 1:
                    cell.alignment = Alignment(horizontal="right")

        for col in range(1, len(headers) + 1):
            max_length = len(headers[col - 1])
            for row in range(2, len(stats) + 2):
                cell_value = sheet.cell(row=row, column=col).value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))
            sheet.column_dimensions[get_column_letter(col)].width = min(max_length + 2, 20)
